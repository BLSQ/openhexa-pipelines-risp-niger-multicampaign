from openhexa.sdk import current_run, parameter, pipeline
import os
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from config import (
    OUTPUTS_PATH,
    TEMPLATES_PATH,
    rename_dict,
    age_group_campaign_dict,
)


@pipeline(
    "generate_targets_templates",
    name="multi-campagne - 01 - Pipeline de création de fichier template pour les cibles",
)
@parameter(
    "campaign",
    name="Campagne",
    help="Sélectionnez le type de campagne",
    type=str,
    required=True,
    choices=[
        "Polio",
        "Rougeole",
        "Méningite",
        "TCV",
        "Fièvre jaune",
        "Albendazole",
        "Vitamine A",
    ],
)
@parameter(
    "campaign_scale",
    name="Échelle de la campagne",
    help="Sélectionnez l'échelle de la campagne",
    type=str,
    required=True,
    choices=[
        "Nationale",
        "Agadez",
        "Diffa",
        "Dosso",
        "Maradi",
        "Niamey",
        "Tahoua",
        "Tillaberi",
        "Zinder",
    ],
    multiple=True,
)
@parameter(
    "year",
    name="Année",
    help="Veuillez entrer l'année de la campagne (2026, 2027, etc.)",
    type=int,
    required=True,
    choices=[
        2026,
        2027,
        2028,
        2029,
        2030,
        2031,
        2032,
        2033,
        2034,
        2035,
        2036,
        2037,
        2038,
        2039,
        2040,
        2041,
        2042,
        2043,
        2044,
        2045,
        2046,
        2047,
        2048,
        2049,
        2050,
    ],
)
@parameter(
    "aggregation_level",
    name="Niveau d'agrégation",
    help="Niveau d'agrégation pour les cibles",
    type=str,
    required=True,
    choices=["CSI", "District"],
)
def generate_targets_templates(
    campaign: str,
    campaign_scale: list,
    year: int,
    aggregation_level: str,
):
    """
    This pipeline generates target templates for each campaign type based on the organizational unit tree
    data from IASO. It retrieves the org unit tree data, cleans it, creates template files for each campaign
    type, and saves the cleaned org unit tree data for future use
    """
    inspect_params(campaign_scale, year)
    iaso_org_unit_tree_df_clean = load_data("iaso_org_unit_tree_clean")
    create_template_files(
        iaso_org_unit_tree_df_clean,
        campaign,
        campaign_scale,
        year,
        aggregation_level,
    )


def inspect_params(
    campaign_scale,
    year,
):
    """
    This function runs checks on the parmater choices for year, round, age, site, and strategy
    """
    current_run.log_info("Vérification des choix des paramètres...")

    # for campaign_scale, if 'Nationale' is present, it cannot coexist with any other choice
    if "Nationale" in campaign_scale and len(campaign_scale) > 1:
        current_run.log_error(
            "Le choix 'Nationale' pour le paramètre 'Échelle de la campagne' ne peut pas être sélectionné avec d'autres choix."
        )
        raise

    # year must be a reasonable integer between 2026 and 2050
    if not isinstance(year, int) or year < 2026 or year > 2050:
        current_run.log_error(
            "Le paramètre 'Année' doit être un entier entre 2026 et 2100."
        )
        raise


def load_data(file_name: str) -> pd.DataFrame:
    """Read a table from the workspace database and return it as a dataframe.

    Parameters
    ----------
    file_name : str
        The name of the file to read from.

    Returns
    -------
    pd.DataFrame
        The dataframe containing the file data.
    """
    current_run.log_info(
        f"Lecture des données de la pyramide des unités organisationnelles depuis le fichier {file_name}..."
    )
    try:
        if not os.path.exists(OUTPUTS_PATH):
            current_run.log_error(f"Le chemin {OUTPUTS_PATH} n'existe pas.")
            raise
        file_to_import = os.path.join(OUTPUTS_PATH, f"{file_name}.parquet")
        df = pd.read_parquet(file_to_import)
        current_run.log_info(
            f"Données de la pyramide des unités organisationnelles chargées avec succès depuis le fichier {file_to_import}"
        )

        return df
    except Exception as e:
        current_run.log_error(
            f"Erreur lors de la lecture des données de la pyramide des unités organisationnelles depuis le fichier {file_to_import}: {e}"
        )
        raise


def create_template_files(
    org_unit_df: pd.DataFrame,
    campaign: str,
    campaign_scale: list,
    year: int,
    aggregation_level: str,
) -> pd.DataFrame:
    """
    Create template files for each organizational unit based on the cleaned org unit tree data.

    Args:
        org_unit_df (pd.DataFrame): DataFrame containing the cleaned org unit tree data.
        campaign (str): The name of the campaign.
        campaign_scale (list): The scale of the campaign.
        year (int): The year of the campaign.
        aggregation_level (str): The level of aggregation for the campaign.

    Returns:
        pd.DataFrame: DataFrame containing the template files for each organizational unit.
    """
    current_run.log_info(
        "Création des fichiers templates pour chaque unité organisationnelle..."
    )
    try:
        # select and rename cols
        cols = rename_dict.keys()
        df = org_unit_df[cols].copy()
        df.insert(0, "Pays", "Niger")
        df = df.rename(columns=rename_dict)

        if aggregation_level == "District":
            if "CSI" in df.columns:
                df = df.drop(columns=["CSI", "Commune"])
            df = df.drop_duplicates()

        df = df.sort_values(by=list(df.columns))

        # create new cols for targets based on age range, site type, and strategy type
        new_cols_list = []
        age_range = age_group_campaign_dict[campaign]
        for age in age_range:
            col_name = f"Cible {age}".strip("_ ")
            if col_name not in df.columns and col_name not in new_cols_list:
                new_cols_list.append(col_name)

        new_columns_df = pd.DataFrame(
            {col: "" for col in new_cols_list}, index=df.index
        )
        df = pd.concat([df, new_columns_df], axis=1)

        # filter by campaign scale if not national
        if "Nationale" not in campaign_scale:
            df = df[df["Région"].isin(campaign_scale)]

        # define file path and name
        os.makedirs(TEMPLATES_PATH, exist_ok=True)
        filename = f"Cibles_{campaign}_{year}_{'_'.join([s.lower() for s in campaign_scale])}_{aggregation_level.lower()}_template.xlsx"
        file_path = os.path.join(TEMPLATES_PATH, filename)

        # Use ExcelWriter to apply formatting
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Template")
            workbook = writer.book
            worksheet = writer.sheets["Template"]

            # set color patterns
            gray_fill = PatternFill(
                start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
            )
            yellow_fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

            # apply auto-filter and freeze top row
            total_cols = len(df.columns)
            total_rows = len(df) + 1
            worksheet.auto_filter.ref = (
                f"A1:{get_column_letter(total_cols)}{total_rows}"
            )
            worksheet.freeze_panes = "A2"
            num_untouchable_cols = (
                df.columns.get_loc(new_cols_list[0]) if new_cols_list else total_cols
            )
            for col_idx in range(1, total_cols + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                column_cells = worksheet[col_letter]

                for cell in column_cells:
                    try:
                        if cell.value:
                            cell_len = len(str(cell.value))
                            if cell_len > max_length:
                                max_length = cell_len
                    except:
                        pass
                adjusted_width = min(max_length * 1.1, 50) + 2
                worksheet.column_dimensions[col_letter].width = adjusted_width
                if col_idx <= num_untouchable_cols:
                    for cell in column_cells:
                        cell.fill = gray_fill
                else:
                    for row_idx, cell in enumerate(column_cells, start=1):
                        cell.fill = gray_fill if row_idx == 1 else yellow_fill

        current_run.add_file_output(file_path)

        current_run.log_info(f"Fichier template généré avec succès: {file_path}")
        return df

    except Exception as e:
        current_run.log_error(f"Erreur lors de la création du fichier template: {e}")
        raise


def save_file(df: pd.DataFrame, file_name: str) -> None:
    """
    Save the cleaned org unit tree data to a parquet file.

    Args:
        df (pd.DataFrame): DataFrame containing the cleaned org unit tree data.
        file_name (str): Name of the file to save the DataFrame as.

    Returns:
        None
    """
    current_run.log_info("Enregistrement des données des unités organisationnelles...")
    try:
        if not os.path.exists(OUTPUTS_PATH):
            os.makedirs(OUTPUTS_PATH)
        file_path = os.path.join(
            OUTPUTS_PATH,
            f"{file_name}.parquet",
        )

        df.to_parquet(
            file_path,
            index=False,
        )
        current_run.log_info(
            f"Données des unités organisationnelles enregistrées avec succès: {file_path}"
        )
    except Exception as e:
        current_run.log_error(f"Erreur lors de l'enregistrement du fichier: {e}")
        raise e


if __name__ == "__main__":
    generate_targets_templates()
